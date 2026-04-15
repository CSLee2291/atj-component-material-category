"""
KPI Tracker
-----------
Takes weekly snapshots of ATJ component MATERIAL_CATEGORY completion progress.
Stores history as JSON for trend charts on the KPI dashboard.

Two KPI scopes:
  1. ALL ATJ components (~80K items)
  2. Phase I items (~7,968 items from Phase_I_PUR_2024_2025.xlsx)

Snapshot = one record per ISO week containing:
  - total ATJ components
  - filled (non-blank MATERIAL_CATEGORY)
  - blank
  - completion %
  - breakdown by lifecycle phase
"""
import os
import json
import logging
import threading
from datetime import datetime, timezone

import pandas as pd

from config import settings

logger = logging.getLogger(__name__)

_KPI_DIR = settings.kpi_data_dir
_SNAPSHOTS_FILE = os.path.join(_KPI_DIR, "snapshots.json")
_PHASE1_SNAPSHOTS_FILE = os.path.join(_KPI_DIR, "phase1_snapshots.json")
_PHASE1_EXCEL = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "Phase_I_PUR_2024_2025.xlsx",
)

# Serialize read-modify-write on the JSON file
_file_lock = threading.Lock()


# ---------------------------------------------------------------------------
# JSON persistence
# ---------------------------------------------------------------------------

def _load_snapshots() -> list[dict]:
    """Load existing snapshots from JSON file."""
    if not os.path.exists(_SNAPSHOTS_FILE):
        return []
    try:
        with open(_SNAPSHOTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        logger.warning("Corrupt snapshots file, starting fresh")
        return []


def _save_snapshots(snapshots: list[dict]) -> None:
    """Write snapshots list to JSON file."""
    os.makedirs(_KPI_DIR, exist_ok=True)
    with open(_SNAPSHOTS_FILE, "w", encoding="utf-8") as f:
        json.dump(snapshots, f, indent=2, ensure_ascii=False)


# ---------------------------------------------------------------------------
# Compute metrics from a DataFrame
# ---------------------------------------------------------------------------

def _compute_snapshot(df: pd.DataFrame) -> dict:
    """
    Compute KPI metrics from the full ATJ components DataFrame.
    Expects columns: Item_Number, LifeCycle_Phase, MATERIAL_CATEGORY.
    """
    now = datetime.now(timezone.utc)
    week_id = now.strftime("%G-W%V")  # ISO year + ISO week number

    total = len(df)

    # Determine filled vs blank
    if "MATERIAL_CATEGORY" in df.columns:
        filled_mask = df["MATERIAL_CATEGORY"].fillna("").str.strip().ne("")
    else:
        filled_mask = pd.Series([False] * total, dtype=bool)

    filled = int(filled_mask.sum())
    blank = total - filled
    pct = round(filled / total * 100, 2) if total > 0 else 0.0

    # Breakdown by lifecycle phase
    by_lifecycle = {}
    if "LifeCycle_Phase" in df.columns and total > 0:
        for phase, group in df.groupby("LifeCycle_Phase", dropna=False):
            phase_name = str(phase) if pd.notna(phase) else "(blank)"
            phase_total = len(group)
            if "MATERIAL_CATEGORY" in group.columns:
                phase_filled = int(group["MATERIAL_CATEGORY"].fillna("").str.strip().ne("").sum())
            else:
                phase_filled = 0
            by_lifecycle[phase_name] = {
                "total": phase_total,
                "filled": phase_filled,
                "blank": phase_total - phase_filled,
                "pct": round(phase_filled / phase_total * 100, 2) if phase_total > 0 else 0.0,
            }

    return {
        "timestamp": now.isoformat(),
        "week": week_id,
        "total": total,
        "filled": filled,
        "blank": blank,
        "completion_pct": pct,
        "by_lifecycle": by_lifecycle,
    }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def take_snapshot() -> dict:
    """
    Fetch current ATJ component data, compute metrics, save snapshot.
    If a snapshot for the current ISO week already exists, overwrites it.
    Returns the new snapshot dict.
    """
    from core.data_fetcher import fetch_all_atj_components

    logger.info("Taking KPI snapshot...")
    df = fetch_all_atj_components()

    if df.empty:
        raise RuntimeError("No ATJ components found — cannot take snapshot")

    snapshot = _compute_snapshot(df)

    with _file_lock:
        snapshots = _load_snapshots()
        # Remove any existing snapshot for same week (overwrite)
        snapshots = [s for s in snapshots if s["week"] != snapshot["week"]]
        snapshots.append(snapshot)
        snapshots.sort(key=lambda s: s["week"])
        _save_snapshots(snapshots)

    logger.info(
        "KPI snapshot saved: week=%s total=%d filled=%d blank=%d pct=%.1f%%",
        snapshot["week"], snapshot["total"], snapshot["filled"],
        snapshot["blank"], snapshot["completion_pct"],
    )
    return snapshot


def get_snapshots() -> list[dict]:
    """Return all historical snapshots (for trend charts)."""
    with _file_lock:
        return _load_snapshots()


def get_latest_snapshot() -> dict | None:
    """Return the most recent snapshot, or None."""
    with _file_lock:
        snapshots = _load_snapshots()
    return snapshots[-1] if snapshots else None


def get_detail_data(filter_mode: str = "all", lifecycle: str | None = None) -> list[dict]:
    """
    Return per-item detail for the dashboard table.
    Each row: Item_Number, Item_Desc, MATERIAL_CATEGORY, LifeCycle_Phase,
              MANUFACTURE_NAME, MFR_PART_NUMBER.

    filter_mode: "all" | "filled" | "blank"
    lifecycle:   optional lifecycle phase filter
    """
    from core.data_fetcher import fetch_all_atj_components, fetch_manufacture_for_items_batched

    df = fetch_all_atj_components()
    if df.empty:
        return []

    # Apply filter
    if filter_mode == "filled":
        df = df[df["MATERIAL_CATEGORY"].fillna("").str.strip().ne("")].reset_index(drop=True)
    elif filter_mode == "blank":
        df = df[df["MATERIAL_CATEGORY"].fillna("").str.strip().eq("")].reset_index(drop=True)

    if lifecycle:
        df = df[df["LifeCycle_Phase"] == lifecycle].reset_index(drop=True)

    if df.empty:
        return []

    # Fetch manufacture data for all items
    item_numbers = df["Item_Number"].tolist()
    mfr_df = fetch_manufacture_for_items_batched(item_numbers)

    if not mfr_df.empty:
        # Keep first manufacturer per item (1:many -> 1:1)
        mfr_dedup = mfr_df.drop_duplicates(subset=["Item_Number"], keep="first")
        df = df.merge(mfr_dedup[["Item_Number", "MANUFACTURE_NAME", "MFR_PART_NUMBER"]],
                       on="Item_Number", how="left")
    else:
        df["MANUFACTURE_NAME"] = ""
        df["MFR_PART_NUMBER"] = ""

    # Select and order columns
    cols = ["Item_Number", "Item_Desc", "MATERIAL_CATEGORY", "LifeCycle_Phase",
            "MANUFACTURE_NAME", "MFR_PART_NUMBER"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""

    return df[cols].fillna("").to_dict(orient="records")


# ===========================================================================
# Phase I KPI
# ===========================================================================

def _load_phase1_items() -> list[str]:
    """Read the Material column from Phase_I_PUR_2024_2025.xlsx."""
    if not os.path.exists(_PHASE1_EXCEL):
        raise FileNotFoundError(f"Phase I Excel not found: {_PHASE1_EXCEL}")
    df = pd.read_excel(_PHASE1_EXCEL, sheet_name="Phase_I_List", usecols=["Material"])
    items = df["Material"].dropna().astype(str).str.strip().tolist()
    items = [i for i in items if i]
    return items


def _load_phase1_snapshots() -> list[dict]:
    if not os.path.exists(_PHASE1_SNAPSHOTS_FILE):
        return []
    try:
        with open(_PHASE1_SNAPSHOTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return []


def _save_phase1_snapshots(snapshots: list[dict]) -> None:
    os.makedirs(_KPI_DIR, exist_ok=True)
    with open(_PHASE1_SNAPSHOTS_FILE, "w", encoding="utf-8") as f:
        json.dump(snapshots, f, indent=2, ensure_ascii=False)


def take_phase1_snapshot() -> dict:
    """
    Take a Phase I KPI snapshot:
    - Read item list from Excel
    - Query Denodo for their current MATERIAL_CATEGORY status
    - Compute filled/blank/pct
    - Save to phase1_snapshots.json (one per ISO week)
    """
    from core.data_fetcher import fetch_items_info

    items = _load_phase1_items()
    logger.info("Phase I KPI: %d items from Excel", len(items))

    # Query Denodo in batches of 500
    frames = []
    for i in range(0, len(items), 500):
        batch = items[i:i + 500]
        df = fetch_items_info(batch)
        if not df.empty:
            frames.append(df)
    df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    now = datetime.now(timezone.utc)
    week_id = now.strftime("%G-W%V")

    total_excel = len(items)
    found = len(df)
    not_found = total_excel - found

    if "MATERIAL_CATEGORY" in df.columns and not df.empty:
        filled_mask = df["MATERIAL_CATEGORY"].fillna("").str.strip().ne("")
        filled = int(filled_mask.sum())
    else:
        filled = 0
    blank = found - filled
    pct = round(filled / total_excel * 100, 2) if total_excel > 0 else 0.0

    # Lifecycle breakdown (of found items)
    by_lifecycle = {}
    if "LifeCycle_Phase" in df.columns and not df.empty:
        for phase, group in df.groupby("LifeCycle_Phase", dropna=False):
            phase_name = str(phase) if pd.notna(phase) else "(blank)"
            phase_total = len(group)
            phase_filled = int(group["MATERIAL_CATEGORY"].fillna("").str.strip().ne("").sum())
            by_lifecycle[phase_name] = {
                "total": phase_total,
                "filled": phase_filled,
                "blank": phase_total - phase_filled,
                "pct": round(phase_filled / phase_total * 100, 2) if phase_total > 0 else 0.0,
            }

    snapshot = {
        "timestamp": now.isoformat(),
        "week": week_id,
        "total_excel": total_excel,
        "found_in_denodo": found,
        "not_found": not_found,
        "total": total_excel,
        "filled": filled,
        "blank": total_excel - filled,  # includes not_found items as blank
        "completion_pct": pct,
        "by_lifecycle": by_lifecycle,
    }

    with _file_lock:
        snapshots = _load_phase1_snapshots()
        snapshots = [s for s in snapshots if s["week"] != week_id]
        snapshots.append(snapshot)
        snapshots.sort(key=lambda s: s["week"])
        _save_phase1_snapshots(snapshots)

    logger.info(
        "Phase I snapshot: week=%s total=%d filled=%d blank=%d pct=%.1f%%",
        week_id, total_excel, filled, total_excel - filled, pct,
    )
    return snapshot


def get_phase1_snapshots() -> list[dict]:
    with _file_lock:
        return _load_phase1_snapshots()


def get_phase1_latest() -> dict | None:
    with _file_lock:
        snaps = _load_phase1_snapshots()
    return snaps[-1] if snaps else None


def get_phase1_detail(filter_mode: str = "all", lifecycle: str | None = None) -> list[dict]:
    """
    Return per-item detail for Phase I items.
    Queries Denodo for current status + manufacture data.
    """
    from core.data_fetcher import fetch_items_info, fetch_manufacture_for_items_batched

    items = _load_phase1_items()
    if not items:
        return []

    # Fetch item info in batches
    frames = []
    for i in range(0, len(items), 500):
        batch = items[i:i + 500]
        df = fetch_items_info(batch)
        if not df.empty:
            frames.append(df)
    df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    if df.empty:
        return []

    # Apply filter
    if filter_mode == "filled":
        df = df[df["MATERIAL_CATEGORY"].fillna("").str.strip().ne("")].reset_index(drop=True)
    elif filter_mode == "blank":
        df = df[df["MATERIAL_CATEGORY"].fillna("").str.strip().eq("")].reset_index(drop=True)

    if lifecycle:
        df = df[df["LifeCycle_Phase"] == lifecycle].reset_index(drop=True)

    if df.empty:
        return []

    # Fetch manufacture data
    item_numbers = df["Item_Number"].tolist()
    mfr_df = fetch_manufacture_for_items_batched(item_numbers)

    if not mfr_df.empty:
        mfr_dedup = mfr_df.drop_duplicates(subset=["Item_Number"], keep="first")
        df = df.merge(mfr_dedup[["Item_Number", "MANUFACTURE_NAME", "MFR_PART_NUMBER"]],
                       on="Item_Number", how="left")
    else:
        df["MANUFACTURE_NAME"] = ""
        df["MFR_PART_NUMBER"] = ""

    cols = ["Item_Number", "Item_Desc", "MATERIAL_CATEGORY", "LifeCycle_Phase",
            "MANUFACTURE_NAME", "MFR_PART_NUMBER"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""

    return df[cols].fillna("").to_dict(orient="records")


def get_phase1_batch_status() -> dict:
    """Return how many Phase I items still need AI categorization (blank MATERIAL_CATEGORY)."""
    items = _load_phase1_items()
    return {"total_items": len(items), "excel_path": _PHASE1_EXCEL}


def write_phase1_results_to_excel(results: list[dict]) -> int:
    """
    Write AI categorization results back to Phase_I_PUR_2024_2025.xlsx.
    Only writes items with High or Medium confidence.
    Adds/updates columns: AI_MATERIAL_CATEGORY, AI_CONFIDENCE, AI_REASON, AI_UPDATED_AT.
    Returns number of items written.
    """
    import openpyxl

    # Filter to high/medium only
    good_results = [r for r in results
                    if r.get("AI_confidence", "").lower() in ("high", "medium")]
    if not good_results:
        return 0

    wb = openpyxl.load_workbook(_PHASE1_EXCEL)
    ws = wb["Phase_I_List"]

    # Find or create AI columns (after existing 5 columns)
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    ai_cols = {
        "AI_MATERIAL_CATEGORY": None,
        "AI_CONFIDENCE": None,
        "AI_REASON": None,
        "AI_UPDATED_AT": None,
    }
    next_col = ws.max_column + 1
    for col_name in ai_cols:
        if col_name in headers:
            ai_cols[col_name] = headers[col_name]
        else:
            ai_cols[col_name] = next_col
            ws.cell(1, next_col, col_name)
            next_col += 1

    # Build a lookup: Item_Number -> row number
    material_col = headers.get("Material", 1)
    item_to_row = {}
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, material_col).value
        if val:
            item_to_row[str(val).strip()] = r

    # Write results
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    written = 0
    for result in good_results:
        item_no = result.get("Item_Number", "")
        row = item_to_row.get(item_no)
        if row is None:
            continue
        ws.cell(row, ai_cols["AI_MATERIAL_CATEGORY"], result.get("AI_MATERIAL_CATEGORY", ""))
        ws.cell(row, ai_cols["AI_CONFIDENCE"], result.get("AI_confidence", ""))
        ws.cell(row, ai_cols["AI_REASON"], result.get("AI_reason", ""))
        ws.cell(row, ai_cols["AI_UPDATED_AT"], now_str)
        written += 1

    wb.save(_PHASE1_EXCEL)
    wb.close()
    logger.info("Wrote %d AI results to Phase I Excel", written)
    return written
