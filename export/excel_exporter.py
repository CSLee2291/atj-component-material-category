"""
Excel Export for CE/PLM ECO Review
------------------------------------
Produces a formatted .xlsx with:
  - Sheet 1: Results (filtered view for CE)
  - Sheet 2: Raw (all columns for debugging)
"""
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import settings

CONFIDENCE_COLORS = {
    "high":   "C6EFCE",  # green
    "medium": "FFEB9C",  # amber
    "low":    "FFC7CE",  # red
    "error":  "D9D9D9",  # gray
}

CE_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")  # blue — CE input cols
HEAD_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=9)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _col_widths():
    return {
        "Item_Number":           18,
        "Item_Desc":             36,
        "LifeCycle_Phase":       16,
        "MANUFACTURE_NAME":      16,
        "MFR_PART_NUMBER":       22,
        "Ref_Item_Number":       18,
        "Ref_MPN":               22,
        "Ref_Similarity":        12,
        "Ref_MATERIAL_CATEGORY": 22,
        "Ref_CATE_M_NAME":       20,
        "Ref_CATE_S_NAME":       20,
        "AI_ZZMCATG_M":          14,
        "AI_ZZMCATG_S":          14,
        "AI_MATERIAL_CATEGORY":  22,
        "AI_confidence":         12,
        "AI_reason":             50,
        "AI_source":             14,
        "Vector_used":           10,
        "Vector_top1_category":  22,
        "Vector_top1_score":     14,
        "CE_MATERIAL_CATEGORY":  22,
        "CE_approved":           12,
        "CE_comment":            30,
        "processed_at":          20,
    }


def export_to_excel(
    df: pd.DataFrame,
    batch_id: str,
    lifecycle_filter: list[str] | None = None,
) -> str:
    os.makedirs(settings.output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    phase_tag = "_".join(lifecycle_filter) if lifecycle_filter else "all_phases"
    filename = f"ATJ_category_{batch_id}_{phase_tag}_{ts}.xlsx"
    filepath = os.path.join(settings.output_dir, filename)

    ce_cols  = ["CE_MATERIAL_CATEGORY", "CE_approved", "CE_comment"]
    col_order = list(_col_widths().keys())
    # Only include columns that exist in df
    col_order = [c for c in col_order if c in df.columns]
    review_df = df[col_order].copy()

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        review_df.to_excel(writer, sheet_name="CE Review", index=False)
        df.to_excel(writer, sheet_name="Raw Data", index=False)

    # Apply formatting
    wb = load_workbook(filepath)

    for sheet_name in ["CE Review", "Raw Data"]:
        ws = wb[sheet_name]
        widths = _col_widths()

        # Header row
        for col_idx, col_name in enumerate(
            review_df.columns if sheet_name == "CE Review" else df.columns, start=1
        ):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEAD_FONT
            cell.fill = HEAD_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 14)

        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"

        # Data rows (CE Review only gets confidence coloring + CE column highlight)
        if sheet_name == "CE Review":
            conf_col = col_order.index("AI_confidence") + 1 if "AI_confidence" in col_order else None
            for row_idx in range(2, ws.max_row + 1):
                confidence = ws.cell(row=row_idx, column=conf_col).value if conf_col else ""
                row_color = CONFIDENCE_COLORS.get(str(confidence).lower(), "FFFFFF")
                conf_fill = PatternFill("solid", start_color=row_color, end_color=row_color)

                for col_idx in range(1, len(col_order) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    col_name = col_order[col_idx - 1]
                    cell.font = BODY_FONT
                    cell.border = BORDER
                    cell.alignment = Alignment(vertical="top", wrap_text=(col_name in ("Item_Desc", "AI_reason")))
                    if col_name in ce_cols:
                        cell.fill = CE_FILL
                    elif col_name == "AI_confidence":
                        cell.fill = conf_fill

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    summary_data = [
        ["Batch ID",           batch_id],
        ["Generated at",       datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total items",        len(df)],
        ["Lifecycle filter",   ", ".join(lifecycle_filter) if lifecycle_filter else "All"],
        ["High confidence",    len(df[df["AI_confidence"] == "high"])],
        ["Medium confidence",  len(df[df["AI_confidence"] == "medium"])],
        ["Low confidence",     len(df[df["AI_confidence"] == "low"])],
        ["Errors",             len(df[df["AI_confidence"] == "error"])],
        ["", ""],
        ["Color legend", ""],
        ["Green row",    "High confidence — AI suggestion reliable"],
        ["Amber row",    "Medium confidence — CE review recommended"],
        ["Red row",      "Low confidence — CE must verify"],
        ["Gray row",     "Error — GPT call failed, manual categorization required"],
        ["Blue columns", "CE input columns — fill in CE_MATERIAL_CATEGORY + CE_approved"],
    ]
    for r, row in enumerate(summary_data, 1):
        for c, val in enumerate(row, 1):
            ws_sum.cell(row=r, column=c, value=val)
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 50

    wb.save(filepath)
    return filepath
